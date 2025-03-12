import openpyxl
import re
import os

# Настройки
output_sheet = "Processed Data"
keywords = [
    'Радіус дії',
    'Радіус',
    'Дальність польоту'
]

# === НОВЫЙ БЛОК ГЕНЕРАЦИИ КЛЮЧЕВЫХ СЛОВ ===
# Создаем расширенный список ключевых слов с разными регистрами
all_keywords = []
for keyword in keywords:
    # Генерируем все варианты
    variations = {
        keyword,  # Оригинальный регистр
        keyword.lower(),
        keyword.upper(),
        keyword.title(),
        ' '.join([w.capitalize() for w in keyword.split()])  # Каждое слово с заглавной
    }
    # Добавляем специфические комбинации для многословных ключей
    if ' ' in keyword:
        words = keyword.split()
        variations.add(words[0].capitalize() + ' ' + ' '.join(words[1:]).lower())
    all_keywords.extend(variations)

# Удаляем дубликаты и обновляем основной список
keywords = list(set(all_keywords))
# =========================================

# Режимы обработки
MODES = {
    1: "Только числа (числовой формат)",
    2: "Текст без ключевых слов",
    3: "Текст без ключевых слов и без символов",
    4: "Текст с ключевыми словами",
    5: "Текст с ключевыми словами без символов",
    6: "Вся строка без символов",
    7: "Вся строка (с символами)"
}

# Функция для извлечения чисел из строки
def extract_numbers(text):
    items = re.split(r'[\n•]', text)
    numbers = []

    for item in items:
        # Теперь проверяем все сгенерированные варианты ключевых слов
        if any(keyword in item for keyword in keywords):
            found = re.findall(r'(\d+([.,]\d+)?)', item)
            nums = [num[0].replace(',', '.') for num in found]
            numbers.extend(nums)

    return numbers


# Функция для очистки текста в зависимости от режима
def clean_text(text, mode):
    if mode in [2, 3]:
        # Удаляем все варианты ключевых слов
        for kw in keywords:
            text = text.replace(kw, '').strip()

    if mode in [3, 5, 6]:
        text = re.sub(r'[^\w\s]', '', text).strip()
    return text.strip()


# Создание начальной таблицы
def create_input_table(file_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Input Data"
    ws.append(["Article", "Original Text"])
    ws.append(["Вставьте ваш текст в столбец B (Original Text)"])
    wb.save(file_path)
    print(f"Таблица создана: {file_path}")


# Обработка данных и создание результирующей таблицы
def process_data(file_path, mode):
    try:
        wb = openpyxl.load_workbook(file_path)
        ws_input = wb["Input Data"]

        # Определяем максимальное количество чисел в строках для режима 1
        max_numbers = 0
        if mode == 1:
            for row in ws_input.iter_rows(min_row=2, max_col=2):
                text = row[1].value
                if text:
                    numbers = extract_numbers(text)
                    max_numbers = max(max_numbers, len(numbers))

        # Создаем лист результатов
        if output_sheet in wb.sheetnames:
            del wb[output_sheet]
        ws_output = wb.create_sheet(output_sheet)

        # Формируем заголовки
        headers = ["Article", "Original Text"]
        if mode == 1:
            headers += [f"Number {i + 1}" for i in range(max_numbers)]
        else:
            headers += ["Processed Data"]
        ws_output.append(headers)

        # Обрабатываем данные
        for row in ws_input.iter_rows(min_row=2, max_col=2):
            article = row[0].value
            text = row[1].value

            # Пропускаем пустые строки
            if not text or not article:
                continue

            output_row = [article, text]

            if mode == 1:  # Только числа
                numbers = extract_numbers(text)
                # Преобразуем числа в строки с запятой вместо точки
                formatted_numbers = [num.replace('.', ',') for num in numbers]
                output_row += formatted_numbers + [''] * (len(headers) - len(output_row) - len(formatted_numbers))
            elif mode in [2, 3, 4, 5, 6, 7]:  # Текстовые режимы
                processed_text = ""
                items = re.split(r'[\n•]', text)
                for item in items:
                    if any(keyword in item for keyword in keywords):
                        if mode == 2:
                            processed_text += clean_text(item, mode) + " "
                        elif mode == 3:
                            processed_text += clean_text(item, mode) + " "
                        elif mode == 4:
                            processed_text += clean_text(item, mode) + " "
                        elif mode == 5:
                            processed_text += clean_text(item, mode) + " "
                        elif mode == 6:
                            processed_text += clean_text(item, mode) + " "
                        elif mode == 7:
                            processed_text += item.strip() + " "
                output_row.append(processed_text.strip())

            # Добавляем строку в таблицу
            ws_output.append(output_row)

        # Сохраняем файл
        wb.save(file_path)
        print(f"Обработка завершена! Результаты сохранены в лист '{output_sheet}'")

    except Exception as e:
        print(f"Ошибка: {e}")


# Основной процесс
def main():
    file_path = os.path.join(os.getcwd(), "Таблиця_вхід.xlsx")

    # Создаем файл, если его нет
    if not os.path.exists(file_path):
        create_input_table(file_path)
        input("Заполните таблицу в Excel и нажмите Enter...")

    # Выбор режима
    print("Выберите режим обработки:")
    for key, value in MODES.items():
        print(f"{key}: {value}")
    try:
        mode = int(input("Введите номер режима: "))
        if mode not in MODES:
            raise ValueError("Неверный режим")
    except ValueError:
        print("Ошибка: Введите корректный номер режима.")
        return

    # Обрабатываем данные
    process_data(file_path, mode)


if __name__ == "__main__":
    main()