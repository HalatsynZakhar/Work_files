import openpyxl
from openpyxl.drawing.image import Image as OpenpyxlImage
from PIL import Image as PILImage
import os
import re
import sys
from pathlib import Path

# Функция для проверки существования пути
def check_path(path, is_folder=False):
    path_obj = Path(path)
    if not path_obj.exists():
        print(f"ОШИБКА: {'Папка' if is_folder else 'Файл'} не существует: {path}")
        sys.exit(1)
    if is_folder and not path_obj.is_dir():
        print(f"ОШИБКА: Указанный путь не является папкой: {path}")
        sys.exit(1)
    if not is_folder and not path_obj.is_file():
        print(f"ОШИБКА: Указанный путь не является файлом: {path}")
        sys.exit(1)

# Путь к папке с изображениями
folder_path = "C:/Users/ABM/Desktop/Image_1c/"
# Путь к существующему файлу Excel
file_path = r"C:\Users\ABM\Desktop\Таблиця хааркеристик.xlsx"
# Проверяем существование путей
check_path(folder_path, is_folder=True)
check_path(file_path, is_folder=False)
# Выбор исходной ячейки (столбец с именами изображений)
input_column = 'B'  # Входной столбец с именами изображений
# Выбор целевой ячейки (столбец, куда будут вставляться изображения)
output_column = 'A'  # Столбец, в который вставляются изображения

# Функция для нормализации имен (удаление пробелов и символов, приведение к нижнему регистру)
def normalize_name(name):
    if name is None:
        return ""
    # Удаляем все символы, кроме букв и цифр, и приводим к нижнему регистру
    return re.sub(r'[^a-zA-Z0-9а-яА-Я]', '', str(name)).lower()

# Функция для удаления фона изображения
def remove_background(image_path, output_path, threshold=240):
    try:
        # Открываем изображение
        img = PILImage.open(image_path).convert("RGBA")
        datas = img.getdata()

        # Создаем новый список пикселей с прозрачным фоном
        new_data = []
        for item in datas:
            # Если пиксель близок к белому, делаем его прозрачным
            if all(value > threshold for value in item[:3]):  # RGB значения выше порога
                new_data.append((255, 255, 255, 0))  # Прозрачный пиксель
            else:
                new_data.append(item)

        # Применяем новые данные к изображению
        img.putdata(new_data)
        img.save(output_path, "PNG")  # Сохраняем во временный файл в формате PNG
        return output_path
    except Exception as e:
        print(f"ОШИБКА при удалении фона изображения {image_path}: {e}")
        return None

# Попытка открыть файл
try:
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active  # Получаем активный лист
except FileNotFoundError:
    print(f"Файл {file_path} не найден. Проверьте правильность пути.")
    sys.exit(1)
except PermissionError:
    print(f"ОШИБКА: Невозможно открыть файл {file_path}.")
    print("Файл может быть открыт в другой программе. Закройте Excel и попробуйте снова.")
    sys.exit(1)
except Exception as e:
    print(f"ОШИБКА при открытии файла: {e}")
    sys.exit(1)

# Начинаем с первой строки
start_row = 2  # Начальная строка для обработки
# Получаем список всех файлов в папке и нормализуем их имена
file_dict = {}
try:
    for filename in os.listdir(folder_path):
        if filename.lower().endswith(('.jpg', '.jpeg', '.png')):
            name_without_ext = os.path.splitext(filename)[0]
            normalized_name = normalize_name(name_without_ext)
            file_dict[normalized_name] = filename
except PermissionError:
    print(f"ОШИБКА: Отказано в доступе к папке {folder_path}.")
    print("Проверьте права доступа и попробуйте снова.")
    sys.exit(1)
except Exception as e:
    print(f"ОШИБКА при чтении содержимого папки: {e}")
    sys.exit(1)

# Проверяем, найдены ли изображения
if not file_dict:
    print(f"ПРЕДУПРЕЖДЕНИЕ: В папке {folder_path} не найдено изображений в формате JPG, JPEG или PNG.")
    choice = input("Хотите продолжить? (y/n): ")
    if choice.lower() != 'y':
        print("Операция отменена пользователем.")
        sys.exit(0)

# Обрабатываем все ячейки в исходной колонке начиная с указанной строки
try:
    for row in range(start_row, ws.max_row + 1):
        cell_value = ws[f'{input_column}{row}'].value  # Получаем значение из исходной ячейки
        if cell_value:  # Если значение не пустое
            # Нормализуем имя из ячейки
            normalized_cell_value = normalize_name(cell_value)
            # Ищем соответствующий файл в словаре нормализованных имен
            if normalized_cell_value in file_dict:
                actual_filename = file_dict[normalized_cell_value]
                image_path = os.path.join(folder_path, actual_filename)
                print(f"Изображение найдено: {image_path}")
                try:
                    # Создаем временный файл
                    temp_image_path = os.path.join(folder_path, f"temp_{actual_filename}")

                    # Удаляем фон изображения во временном файле
                    processed_image_path = remove_background(image_path, temp_image_path)

                    if processed_image_path:
                        # Загружаем обработанное изображение
                        pil_img = PILImage.open(processed_image_path)
                        img_width, img_height = pil_img.size

                        # Создаем объект Image для добавления в Excel
                        img = OpenpyxlImage(processed_image_path)

                        # Получаем размеры целевой ячейки
                        cell_width = ws.column_dimensions[output_column].width  # Получаем ширину целевого столбца
                        if cell_width is None:
                            cell_width = 10  # Устанавливаем значение по умолчанию
                        cell_height = ws.row_dimensions[row].height  # Получаем высоту строки
                        if cell_height is None:
                            cell_height = 15  # Устанавливаем значение по умолчанию

                        # Размеры ячейки в пикселях (коэффициенты для преобразования)
                        cell_width_px = cell_width * 7  # Преобразуем ширину в пиксели
                        cell_height_px = cell_height * 1.5  # Преобразуем высоту в пиксели

                        # Сохраняем пропорции: подгоняем изображение, чтобы оно вписалось в ячейку
                        scale_factor = min(cell_width_px / img_width, cell_height_px / img_height)
                        img.width = int(img_width * scale_factor)
                        img.height = int(img_height * scale_factor)

                        # Привязываем изображение к ячейке
                        img.anchor = f'{output_column}{row}'  # Размещение изображения в целевой ячейке

                        # Центрируем изображение в ячейке
                        img.left = (cell_width_px - img.width) / 2  # Центрируем по горизонтали
                        img.top = (cell_height_px - img.height) / 2  # Центрируем по вертикали

                        # Вставляем изображение в ячейку
                        ws.add_image(img, f'{output_column}{row}')

                        # Удаляем временный файл после использования
                        os.remove(temp_image_path)
                except PermissionError:
                    print(f"ОШИБКА: Отказано в доступе к изображению: {image_path}")
                    print("Проверьте права доступа к файлу.")
                except Exception as e:
                    print(f"ОШИБКА при обработке изображения {image_path}: {e}")
            else:
                print(f"Изображение для '{cell_value}' не найдено в папке: {folder_path}")
        else:
            print(f"В ячейке {input_column}{row} нет имени изображения.")
    # Сохраняем файл Excel
    try:
        wb.save(file_path)
        print(f"Файл успешно сохранен: {file_path}")
    except PermissionError:
        print(f"ОШИБКА: Невозможно сохранить файл {file_path}.")
        print("Файл может быть открыт в другой программе. Закройте Excel и попробуйте снова.")
        sys.exit(1)
    except Exception as e:
        print(f"ОШИБКА при сохранении файла: {e}")
        sys.exit(1)
except Exception as e:
    print(f"ОШИБКА при выполнении программы: {e}")
    sys.exit(1)
print("Работа программы завершена.")